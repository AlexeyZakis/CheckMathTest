import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl import load_workbook
from config import *

wb = load_workbook(f'C:/Users/{UserName}/Desktop/{FileName}.xlsx')
Active = wb.active
Active.title = SheetName

Alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
			"M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X",
			"Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH",
			"AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR",
			"AS", "AT", "AU", "AV", "AW", "AX", "AY"]


def CheckMaxRow():
	MaxRow = 0
	for i in Active["A"]:
		if i.value == None: break
		MaxRow += 1
	return MaxRow


def CheckMaxCol():
	MaxCol = 0
	for i in Active[1]:
		if i.value == None: break
		MaxCol += 1
	return MaxCol


def TableStat(MaxRow, MaxCol):
	Data = []
	NumRow = 2
	for row in Active.iter_rows(min_row=2, min_col=5, max_col=MaxCol, max_row=MaxRow):
		Dict = {
				"Name" : Active[f"A{NumRow}"].value,
				"Class" : Active[f"B{NumRow}"].value,
				"NumCorrectAnswers" : Active[f"C{NumRow}"].value,
				"Mark" : Active[f"D{NumRow}"].value
				}
		NumTask = 1
		for cell in row:
			Dict[f"Task_{NumTask}"] = cell.value
			NumTask += 1
		NumRow += 1
		Data.append(Dict)
	return Data


def TableSort(MaxRow, MaxCol):
	Data = TableStat(MaxRow=MaxRow, MaxCol=MaxCol)
	NumRow = 2

	for NumCorrectAnswers in reversed(range(0, MaxScore+1)):
		for Student in Data:
			if Student["NumCorrectAnswers"] == str(NumCorrectAnswers):

				Active[f"A{NumRow}"].value = Student["Name"]
				Active[f"B{NumRow}"].value = Student["Class"]
				Active[f"C{NumRow}"].value = Student["NumCorrectAnswers"]
				Active[f"D{NumRow}"].value = Student["Mark"]

				NumTask = 1
				NumLetter = 4

				for col in Active.iter_cols(min_row=NumRow, min_col=5, max_col=MaxCol, max_row=NumRow):
					for cell in col:
						Active[f"{Alphabet[NumLetter]}{NumRow}"].value = Student[f"Task_{NumTask}"]
						NumLetter += 1
						NumTask += 1
				NumRow += 1


def SetBorder(MaxCol, MaxRow):
	border = Border(left=Side(border_style='thin', color=CellBorderColor),
					right=Side(border_style='thin', color=CellBorderColor),
					top=Side(border_style='thin', color=CellBorderColor),
					bottom=Side(border_style='thin', color=CellBorderColor))

	for col in Active.iter_cols(min_row=1, min_col=1, max_col=MaxCol, max_row=MaxRow):
		for cell in col:
			cell.border = border


def ChangeColWidth(MaxCol, ColWidth, MinCol=2, NumLetter=1):
	for col in Active.iter_cols(min_col=MinCol, max_col=MaxCol):
		Active.column_dimensions[Alphabet[NumLetter]].width = ColWidth
		NumLetter += 1


def ChangeRowHeight(NumRow, RowHeight): 
	Active.row_dimensions[NumRow].height = RowHeight


def ChangeAlignment(MaxCol):
	for row in Active.iter_rows(min_row=2, min_col=1, max_col=MaxCol, max_row=2):
		for cell in row:
			cell.alignment = Alignment(horizontal='general')


def DelCol(MinRow, MaxRow, MinCol, MaxCol, NumLetter):
	NumRow = 0
	for col in Active.iter_cols(min_row=MinRow, min_col=MinCol, max_col=MaxCol, max_row=MaxRow+1):
		for cell in col:
			NumRow += 1
			Active[f"{Alphabet[NumLetter]}{NumRow}"] = None
		NumRow = 0
		NumLetter += 1


def DelFontBold(MaxCol):
	font = Font(bold=False)
	for row in Active.iter_rows(min_row=2, min_col=1, max_col=MaxCol, max_row=2):
		for cell in row:
			cell.font = font


def DelBGC(MinCol, MaxCol, MinRow, MaxRow):
	fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
	for row in Active.iter_rows(min_row=MinRow, min_col=MinCol, max_col=MaxCol, max_row=MaxRow):
		for cell in row:
			cell.fill = fill


def UnmergeCells(MaxCol):
	for Col in range(1, MaxCol+1):
		Active.unmerge_cells(start_row=1, start_column=Col, end_row=2, end_column=Col)


def MoveRow(MinRow, MaxRow, MaxCol, NumRow):
	NumLetter = 0
	for row in Active.iter_rows(min_row=MinRow, min_col=1, max_col=MaxCol, max_row=MaxRow + (MinRow-1)):
		for cell in row:
			Active[f"{Alphabet[NumLetter]}{NumRow}"] = cell.value
			NumLetter += 1
		NumLetter = 0
		NumRow += 1


def MoveCol(MinCol, MaxCol, MaxRow, NumLetter):
	NumRow = 0
	for col in Active.iter_cols(min_row=1, min_col=MinCol, max_col=MaxCol, max_row=MaxRow):
		for cell in col:
			NumRow += 1
			Active[f"{Alphabet[NumLetter]}{NumRow}"] = str(cell.value)
		NumRow = 0
		NumLetter += 1


def SetRightAnswers(MaxCol):
	Answers = []
	if Active["C2"].value == str(MaxScore):
		for col in Active.iter_cols(min_row=2, max_row=2, min_col=5, max_col=MaxCol):
			for cell in col:
				Answers.append(cell.value)
	else:
		Answers = [Answer1, Answer2, Answer3, Answer4, Answer5, Answer6, Answer7,
				   Answer8, Answer9, Answer10, Answer11, Answer12, Answer13,
				   Answer14, Answer15, Answer16, Answer17, Answer18, Answer19,
				   Answer20, Answer21, Answer22, Answer23, Answer24, Answer25,
				   Answer26, Answer27, Answer28, Answer29, Answer30, Answer31,
				   Answer32, Answer33, Answer34, Answer35, Answer36, Answer37,
				   Answer38, Answer39, Answer40]
	return Answers


def CountMistakes(MaxRow, MaxCol):
	Active.merge_cells(f'A{MaxRow+1}:D{MaxRow+1}')

	Name = Active[f"A{MaxRow+1}"]
	Name.alignment = Alignment(horizontal='center')
	Name.value = "Количество ошибок"
	Name.font = Font(bold=True, size=14)

	Answers = SetRightAnswers(MaxCol=MaxCol)
	NumCol = 0
	NumMistakes = 0

	for col in Active.iter_cols(min_row=2, min_col=5, max_col=MaxCol, max_row=MaxRow):
		for cell in col:
			if cell.value != Answers[NumCol]:
				NumMistakes += 1
		Active[f"{Alphabet[cell.column-1]}{cell.row+1}"].value = str(NumMistakes)
		NumCol += 1
		NumMistakes = 0

	MaxRow = CheckMaxRow()

	for row in Active.iter_rows(min_row=MaxRow, min_col=5, max_col=MaxCol, max_row=MaxRow):
		for cell in row:
			cell.font = Font(bold=True, size=14)
			if int(cell.value) / (MaxRow - 2) >= CriticalNumMistakes:
				cell.fill = PatternFill(fill_type='solid', start_color=MistakesColor)
				Active[f"{Alphabet[cell.column-1]}1"].fill = PatternFill(fill_type='solid', start_color=MistakesColor)


def ColorMistakes(MaxCol, MaxRow):
	Answers = SetRightAnswers(MaxCol=MaxCol)

	NumAnswer = 0
	for col in Active.iter_cols(min_row=2, min_col=5, max_col=MaxCol, max_row=MaxRow):
		for cell in col:
			if cell.value != str(Answers[NumAnswer]) and cell.value != None:
				cell.fill = PatternFill(fill_type='solid', start_color=MistakesColor)
		NumAnswer += 1


def FormatTable():
	MaxCol = CheckMaxCol()

	UnmergeCells(MaxCol=11)

	Active["A2"].value = "1"
	MaxRow = CheckMaxRow()
	
	MoveRow(MinRow=3, MaxRow=MaxRow, MaxCol=MaxCol, NumRow=2)

	MaxRow = CheckMaxRow()

	MoveCol(MinCol=7, MaxCol=9, MaxRow=MaxRow, NumLetter=0)
	MoveCol(MinCol=11, MaxCol=MaxCol, MaxRow=MaxRow, NumLetter=3)
	DelCol(MinRow=1, MaxRow=MaxRow, MinCol=MaxCol-7, MaxCol=MaxCol, NumLetter=MaxCol-7)

	MaxCol = CheckMaxCol()
	ChangeColWidth(NumLetter=0, MinCol=1, MaxCol=1, ColWidth=NameColWidth)
	ChangeColWidth(MaxCol=MaxCol, ColWidth=ColWidth)

	SetBorder(MaxCol=MaxCol, MaxRow=MaxRow)
	TableSort(MaxCol=MaxCol, MaxRow=MaxRow)
	DelBGC(MinCol=1, MaxCol=MaxCol, MinRow=2, MaxRow=2)
	DelBGC(MinCol=MaxCol+1, MaxCol=MaxCol+8, MinRow=1, MaxRow=2)
	DelFontBold(MaxCol=MaxCol)
	ChangeAlignment(MaxCol=MaxCol)
	ChangeRowHeight(NumRow=1, RowHeight=TitleRowHeight)
	ChangeRowHeight(NumRow=2, RowHeight=BaseRowHeight)
	ColorMistakes(MaxCol=MaxCol, MaxRow=MaxRow)
	CountMistakes(MaxRow=MaxRow, MaxCol=MaxCol)

FormatTable()

wb.save(f'C:/Users/{UserName}/Desktop/{FileName}.xlsx')